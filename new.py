# Importing Libraries
from PIL import Image
from PIL.ExifTags import TAGS
import glob
import os
from openpyxl import Workbook, load_workbook
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl import formatting, styles

# Taking all images files from folder
images = glob.glob("*.jpg")
result = []
for image in images:
    # Path to the image 
    imagename = image
    absolute_path = os.path.abspath(image)
    # Read the image data using PIL
    image = Image.open(imagename)
    # Extracting data
    info_dict = [image.filename,image.format,absolute_path,str(image.size), str(image.height), str(image.width)]
    result.append(info_dict)

# Saving to excel
wb = openpyxl.Workbook()
ws1 = wb.active
# Inserting headers
headers = ["Image Name", "Image Extension", "Full Path", "Image Size", "Image Height", "Image Width"]  
ws1.append(headers)
# Inserting values 
for row in result:   
    ws1.append(row)
#wb.save("Images_Data.xlsx")

light_green_font = Font(color='00FF00', italic=True)
# Changing the font and color of headers
for cell in ws1["1:1"]:
    cell.font = light_green_font
#wb.save("Images_Data.xlsx")
# Formatting
redfill = PatternFill(start_color='EE4636', end_color='EE4636', fill_type='solid')
#ws1.conditional_formatting.add('E2:E20', CellIsRule(operator='greaterThan', formula=['1000'], stopIfTrue=True, fill=redfill))
#ws1.conditional_formatting.add('F2:F20', CellIsRule(operator='greaterThan', formula=['1000'], stopIfTrue=True, fill=redfill))
ws1.conditional_formatting.add('D2:D20', CellIsRule(operator='greaterThan', formula=['1000'], stopIfTrue=True, fill=redfill))
wb.save("Images_Data.xlsx")